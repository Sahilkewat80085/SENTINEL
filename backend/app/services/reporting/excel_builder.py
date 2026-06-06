"""
Module 10: Reporting Engine - Excel Builder
Generates a 9-sheet Excel workbook using openpyxl with full governance analysis.
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import logger
from app.services.governance_score import GovernanceScoreService
from app.services.folder_health import FolderHealthService
from app.services.folder_coverage import FolderCoverageService
from app.services.content_verification import ContentVerificationService
from app.services.merge_delay import MergeDelayService
from app.services.exception_detection import ExceptionDetectionService
from app.services.trend_analytics import TrendAnalyticsService
from app.repositories import commit_repo, repository_repo

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl is not installed – Excel reporting disabled")


# ---------------------------------------------------------------------------
# Palette constants
# ---------------------------------------------------------------------------
DARK_NAVY = "0D1526"
BLUE = "3B82F6"
EMERALD = "10B981"
AMBER = "F59E0B"
ROSE = "F43F5E"
GRAY_800 = "1E293B"
GRAY_700 = "334155"
WHITE = "FFFFFF"
LIGHT_GRAY = "F1F5F9"


def _make_font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)


def _make_fill(hex_color: str):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _make_border(color="475569"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_row(ws, headers: list[str], row: int = 1):
    """Write a styled header row to a worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _make_font(bold=True, color=WHITE, size=10)
        cell.fill = _make_fill(BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _make_border()


def _data_row(ws, values: list[Any], row: int, alt: bool = False):
    """Write a styled data row to a worksheet."""
    bg = "131E31" if alt else "0D1526"
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _make_font(color=LIGHT_GRAY, size=10)
        cell.fill = _make_fill(bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = _make_border()


def _add_title_block(ws, title: str, subtitle: str, generated_at: datetime):
    """Add a rich title block at the top of a worksheet."""
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"🛡 SENTINEL — {title}"
    title_cell.font = Font(bold=True, color=WHITE, size=14)
    title_cell.fill = _make_fill(DARK_NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:I2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"{subtitle} | Generated: {generated_at.strftime('%Y-%m-%d %H:%M UTC')}"
    subtitle_cell.font = Font(color="64748B", size=9, italic=True)
    subtitle_cell.fill = _make_fill(DARK_NAVY)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16


def _color_severity(cell, severity: str):
    """Apply severity-specific fill and font to a cell."""
    color_map = {
        "CRITICAL": (ROSE, WHITE),
        "HIGH": (AMBER, "0D1526"),
        "MEDIUM": ("7C3AED", WHITE),
        "LOW": ("0EA5E9", WHITE),
    }
    bg, fg = color_map.get(severity.upper(), (GRAY_800, WHITE))
    cell.fill = _make_fill(bg)
    cell.font = _make_font(bold=True, color=fg, size=9)


def _color_status(cell, status: str):
    """Apply status-specific fill and font to a cell."""
    color_map = {
        "EXCELLENT": (EMERALD, WHITE),
        "GOOD": ("22C55E", WHITE),
        "WARNING": (AMBER, "0D1526"),
        "POOR": ("F97316", WHITE),
        "CRITICAL": (ROSE, WHITE),
        "MERGED": (EMERALD, WHITE),
        "PARTIAL": (AMBER, "0D1526"),
        "MISSING": (ROSE, WHITE),
        "HEALTHY": (EMERALD, WHITE),
    }
    bg, fg = color_map.get(status.upper(), (GRAY_800, WHITE))
    cell.fill = _make_fill(bg)
    cell.font = _make_font(bold=True, color=fg, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


class ExcelReportBuilder:
    """Builds a 9-sheet Excel governance workbook."""

    def __init__(self):
        self.governance_score_service = GovernanceScoreService()
        self.folder_health_service = FolderHealthService()
        self.coverage_service = FolderCoverageService()
        self.content_service = ContentVerificationService()
        self.delay_service = MergeDelayService()
        self.violation_service = ExceptionDetectionService()
        self.trend_service = TrendAnalyticsService()

    async def build(
        self,
        db: AsyncSession,
        repository_id: uuid.UUID,
        config: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Generate the full workbook and return raw bytes (.xlsx)."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl library not installed. Install with: pip install openpyxl")

        config = config or {}
        now = datetime.now(timezone.utc)
        wb = openpyxl.Workbook()

        repo = await repository_repo.get(db, repository_id)
        repo_name = repo.name if repo else str(repository_id)

        # Remove the default empty sheet that openpyxl creates
        wb.remove(wb.active)

        # --- Sheet 1: Governance Summary ---
        await self._build_summary_sheet(wb, db, repository_id, repo_name, now)

        # --- Sheet 2: Folder Health ---
        await self._build_folder_health_sheet(wb, db, repository_id, now)

        # --- Sheet 3: Coverage Matrix ---
        await self._build_coverage_sheet(wb, db, repository_id, now)

        # --- Sheet 4: Merge Delay Analysis ---
        await self._build_delay_sheet(wb, db, repository_id, now)

        # --- Sheet 5: Content Drift ---
        await self._build_content_drift_sheet(wb, db, repository_id, now)

        # --- Sheet 6: Violation Log ---
        await self._build_violations_sheet(wb, db, repository_id, now)

        # --- Sheet 7: Trend Snapshots ---
        await self._build_trends_sheet(wb, db, repository_id, now)

        # --- Sheet 8: Commit Activity ---
        await self._build_commits_sheet(wb, db, repository_id, now)

        # --- Sheet 9: Audit Trail ---
        self._build_audit_sheet(wb, repo_name, now, config)

        # Write to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    # -----------------------------------------------------------------------
    # Sheet builders
    # -----------------------------------------------------------------------

    async def _build_summary_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, repo_name: str, now: datetime
    ):
        ws = wb.create_sheet("📊 Governance Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

        # Title block
        _add_title_block(ws, "Governance Summary", f"Repository: {repo_name}", now)

        # Score section
        score_res = await self.governance_score_service.compute_repository_score(db, repo_id)
        score_data = score_res.value if score_res.is_success else None

        metrics = []
        if score_data:
            metrics = [
                ("Governance Score", f"{score_data.score}/100"),
                ("Grade", score_data.grade),
                ("Folder Health Average", f"{score_data.folder_health_average:.1f}%"),
                ("Violation Penalty", f"-{score_data.violation_penalty:.1f} pts"),
                ("Critical Violations", str(score_data.active_critical_count)),
                ("High Violations", str(score_data.active_high_count)),
                ("Medium Violations", str(score_data.active_medium_count)),
                ("Low Violations", str(score_data.active_low_count)),
            ]

        _header_row(ws, ["Metric", "Value"], row=4)
        for idx, (metric, value) in enumerate(metrics, start=5):
            _data_row(ws, [metric, value], row=idx, alt=idx % 2 == 0)

        # Apply color to grade cell
        if score_data and len(metrics) >= 2:
            grade_cell = ws.cell(row=6, column=2)
            grade_color = {
                "A": EMERALD, "B": "22C55E", "C": AMBER, "D": "F97316", "E": ROSE, "F": ROSE
            }.get(score_data.grade, GRAY_700)
            grade_cell.fill = _make_fill(grade_color)
            grade_cell.font = _make_font(bold=True, color=WHITE, size=14)
            grade_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[4].height = 20

    async def _build_folder_health_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, now: datetime
    ):
        ws = wb.create_sheet("🏥 Folder Health")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Folder Health Rankings", "Weighted multi-metric health scores per deployment folder", now)

        headers = ["Rank", "Folder", "Health Score", "Coverage", "Consistency", "Timeliness", "Completeness", "Classification"]
        _header_row(ws, headers, row=4)

        col_widths = [8, 30, 15, 14, 14, 14, 14, 16]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        health_res = await self.folder_health_service.compute_all_health(db, repo_id)
        health_data = sorted(health_res.value or [], key=lambda x: x.health_score, reverse=True)

        for rank_idx, fh in enumerate(health_data, start=1):
            row = rank_idx + 4
            values = [
                rank_idx,
                fh.folder_name,
                f"{fh.health_score:.1f}%",
                f"{fh.coverage_score:.1f}%",
                f"{fh.consistency_score:.1f}%",
                f"{fh.timeliness_score:.1f}%",
                f"{fh.completeness_score:.1f}%",
                fh.classification,
            ]
            _data_row(ws, values, row=row, alt=rank_idx % 2 == 0)
            # Color classification cell
            cls_cell = ws.cell(row=row, column=8)
            _color_status(cls_cell, fh.classification)

        ws.row_dimensions[4].height = 20
        ws.freeze_panes = "A5"

    async def _build_coverage_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, now: datetime
    ):
        ws = wb.create_sheet("📋 Coverage Matrix")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Jira Coverage Matrix", "Folder merge status per Jira ticket", now)

        # Get summary + matrix
        summary_res = await self.coverage_service.get_coverage_summary(db, repo_id)
        matrix_res = await self.coverage_service.get_coverage_matrix(db, repo_id)
        summary = summary_res.value if summary_res.is_success else None
        matrix = matrix_res.value if matrix_res.is_success else None

        # Summary stats
        if summary:
            ws.merge_cells("A4:D4")
            sum_cell = ws["A4"]
            sum_cell.value = (
                f"Total: {summary.total_jiras} Jiras | Merged: {summary.merged_count} | "
                f"Partial: {summary.partial_count} | Missing: {summary.missing_count} | "
                f"Coverage: {summary.overall_coverage_pct:.1f}%"
            )
            sum_cell.font = _make_font(color=LIGHT_GRAY, size=10)
            sum_cell.fill = _make_fill(GRAY_800)

        if not matrix or not matrix.rows:
            ws["A6"] = "No coverage data available"
            return

        folders = matrix.folders_list
        headers = ["Jira ID", "Coverage %", "Status"] + folders
        _header_row(ws, headers, row=6)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        for i, folder in enumerate(folders, start=4):
            ws.column_dimensions[get_column_letter(i)].width = 18

        for row_idx, coverage_row in enumerate(matrix.rows, start=7):
            alt = row_idx % 2 == 0
            # Find folder values
            folder_map = {fd.folder_name: ("✓" if fd.is_merged else "✗") for fd in coverage_row.folders}
            values = [
                coverage_row.jira_id,
                f"{coverage_row.coverage_pct:.1f}%",
                coverage_row.status,
            ] + [folder_map.get(f, "—") for f in folders]

            _data_row(ws, values, row=row_idx, alt=alt)
            # Color status
            _color_status(ws.cell(row=row_idx, column=3), coverage_row.status)
            # Color folder cells
            for col_offset, folder in enumerate(folders):
                col_idx = col_offset + 4
                cell = ws.cell(row=row_idx, column=col_idx)
                val = folder_map.get(folder, "—")
                if val == "✓":
                    cell.fill = _make_fill("064E3B")
                    cell.font = _make_font(color=EMERALD, bold=True)
                elif val == "✗":
                    cell.fill = _make_fill("450A0A")
                    cell.font = _make_font(color=ROSE, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A7"

    async def _build_delay_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, now: datetime
    ):
        ws = wb.create_sheet("⏱ Merge Delays")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Merge Delay Analysis", "Propagation delay rankings by deployment folder", now)

        stats_res = await self.delay_service.get_delay_statistics(db, repo_id)
        stats = stats_res.value if stats_res.is_success else None

        if stats:
            ws.merge_cells("A4:F4")
            cell = ws["A4"]
            cell.value = (
                f"Overall Avg: {stats.overall_avg_delay_days:.1f} days | "
                f"Max: {stats.overall_max_delay_days:.1f} days"
            )
            cell.font = _make_font(color=LIGHT_GRAY, size=10)
            cell.fill = _make_fill(GRAY_800)

        headers = ["Rank", "Folder Name", "Avg Delay (days)", "Max Delay (days)", "P95 Delay (days)"]
        _header_row(ws, headers, row=6)

        col_widths = [8, 30, 20, 20, 20]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        if stats and stats.folder_rankings:
            for rank_idx, folder_rank in enumerate(stats.folder_rankings, start=1):
                row = rank_idx + 6
                values = [
                    rank_idx,
                    folder_rank.folder_name,
                    f"{folder_rank.avg_delay_days:.2f}",
                    f"{folder_rank.max_delay_days:.2f}",
                    f"{folder_rank.p95_delay_days:.2f}",
                ]
                _data_row(ws, values, row=row, alt=rank_idx % 2 == 0)
                # Color delay values
                avg_cell = ws.cell(row=row, column=3)
                if folder_rank.avg_delay_days > 14:
                    avg_cell.fill = _make_fill("450A0A")
                    avg_cell.font = _make_font(color=ROSE)
                elif folder_rank.avg_delay_days > 7:
                    avg_cell.fill = _make_fill("451A03")
                    avg_cell.font = _make_font(color=AMBER)
                else:
                    avg_cell.fill = _make_fill("052E16")
                    avg_cell.font = _make_font(color=EMERALD)

        ws.freeze_panes = "A7"

    async def _build_content_drift_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, now: datetime
    ):
        ws = wb.create_sheet("🔍 Content Drift")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Content Drift Report", "SHA256 file hash verification across deployment folders", now)

        drift_res = await self.content_service.get_drift_report(db, repo_id)
        drift = drift_res.value if drift_res.is_success else None

        if drift:
            ws.merge_cells("A4:F4")
            cell = ws["A4"]
            cell.value = f"Drifted Files: {len(drift.drifted_files)} | Overall Drift Score: {drift.overall_drift_score:.2f}"
            cell.font = _make_font(color=LIGHT_GRAY, size=10)
            cell.fill = _make_fill(GRAY_800)

        headers = ["File Path", "Status", "Drift Score", "Divergent Folders", "Majority Hash (Prefix)"]
        _header_row(ws, headers, row=6)

        col_widths = [50, 14, 15, 40, 20]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        if drift:
            for row_idx, result in enumerate(drift.drifted_files, start=7):
                alt = row_idx % 2 == 0
                div_folders = ", ".join(result.divergent_folders) if result.divergent_folders else "None"
                hash_prefix = (result.majority_hash or "")[:12] + "..." if result.majority_hash else "—"
                values = [
                    result.file_path,
                    result.status,
                    f"{result.drift_score:.3f}",
                    div_folders,
                    hash_prefix,
                ]
                _data_row(ws, values, row=row_idx, alt=alt)
                _color_status(ws.cell(row=row_idx, column=2), result.status)

        ws.freeze_panes = "A7"

    async def _build_violations_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, now: datetime
    ):
        ws = wb.create_sheet("⚠️ Violation Log")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Governance Violation Log", "All detected rule violations with acknowledgement status", now)

        violations_res = await self.violation_service.get_violations(db, repository_id=repo_id)
        violations = violations_res.value if violations_res.is_success else []

        summary_res = await self.violation_service.get_violation_summary(db, repo_id)
        summary = summary_res.value if summary_res.is_success else None

        if summary:
            ws.merge_cells("A4:H4")
            cell = ws["A4"]
            cell.value = (
                f"Total: {summary.total_violations} | Critical: {summary.critical_count} | "
                f"High: {summary.high_count} | Medium: {summary.medium_count} | "
                f"Low: {summary.low_count} | Acknowledged: {summary.acknowledged_count}"
            )
            cell.font = _make_font(color=LIGHT_GRAY, size=10)
            cell.fill = _make_fill(GRAY_800)

        headers = ["Rule ID", "Severity", "Category", "Jira", "Folder", "Description", "Detected At", "Status"]
        _header_row(ws, headers, row=6)

        col_widths = [20, 12, 15, 18, 20, 50, 22, 18]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for row_idx, v in enumerate(violations, start=7):
            alt = row_idx % 2 == 0
            detected_str = v.detected_at.strftime("%Y-%m-%d %H:%M") if hasattr(v.detected_at, 'strftime') else str(v.detected_at)
            status_str = "Acknowledged" if v.is_acknowledged else "Active"
            values = [
                v.rule_id,
                v.severity,
                v.category,
                v.jira_id or "—",
                v.folder_name or "—",
                v.description,
                detected_str,
                status_str,
            ]
            _data_row(ws, values, row=row_idx, alt=alt)
            _color_severity(ws.cell(row=row_idx, column=2), v.severity)
            # Status color
            status_cell = ws.cell(row=row_idx, column=8)
            if v.is_acknowledged:
                status_cell.fill = _make_fill("052E16")
                status_cell.font = _make_font(color=EMERALD, bold=True)
            else:
                status_cell.fill = _make_fill("450A0A")
                status_cell.font = _make_font(color=ROSE, bold=True)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A7"

    async def _build_trends_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, now: datetime
    ):
        ws = wb.create_sheet("📈 Trend Snapshots")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Historical Trend Snapshots", "30-day coverage, health, delay, and violation trends", now)

        coverage_res = await self.trend_service.get_coverage_trend(db, repo_id, "30d")
        health_res = await self.trend_service.get_health_trend(db, repo_id, "30d")
        delay_res = await self.trend_service.get_delay_trend(db, repo_id, "30d")
        violations_res = await self.trend_service.get_violation_trend(db, repo_id, "30d")

        coverage_pts = coverage_res.value or []
        health_pts = health_res.value or []
        delay_pts = delay_res.value or []
        violation_pts = violations_res.value or []

        headers = ["Date", "Coverage %", "Health Score", "Avg Delay (days)", "Active Violations"]
        _header_row(ws, headers, row=4)
        col_widths = [22, 15, 15, 20, 20]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        cov_map = {p.date: p.value for p in coverage_pts}
        health_map = {p.date: p.value for p in health_pts}
        delay_map = {p.date: p.value for p in delay_pts}
        viol_map = {p.date: p.value for p in violation_pts}

        all_dates = sorted(set(cov_map) | set(health_map) | set(delay_map) | set(viol_map))

        for row_idx, date in enumerate(all_dates, start=5):
            alt = row_idx % 2 == 0
            values = [
                date,
                f"{cov_map.get(date, 0):.1f}%",
                f"{health_map.get(date, 0):.1f}%",
                f"{delay_map.get(date, 0):.1f}",
                str(int(viol_map.get(date, 0))),
            ]
            _data_row(ws, values, row=row_idx, alt=alt)

        ws.freeze_panes = "A5"

    async def _build_commits_sheet(
        self, wb, db: AsyncSession, repo_id: uuid.UUID, now: datetime
    ):
        ws = wb.create_sheet("💾 Commit Activity")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Commit Activity Log", "Recent 200 commits with author and folder attribution", now)

        # Fetch recent commits - pass limit as keyword arg
        commits_res = await commit_repo.get_commits_for_repository(db, repo_id, limit=200)
        commits = commits_res if commits_res else []

        headers = ["SHA (short)", "Author", "Branch", "Commit Date", "Message"]
        _header_row(ws, headers, row=4)
        col_widths = [12, 25, 20, 22, 80]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for row_idx, commit in enumerate(commits, start=5):
            alt = row_idx % 2 == 0
            author_name = commit.author.name if commit.author else "Unknown"
            commit_date_str = (
                commit.commit_date.strftime("%Y-%m-%d %H:%M")
                if hasattr(commit.commit_date, 'strftime') else str(commit.commit_date)
            )
            message_short = (commit.message or "")[:120]
            values = [
                commit.sha[:8],
                author_name,
                commit.branch or "—",
                commit_date_str,
                message_short,
            ]
            _data_row(ws, values, row=row_idx, alt=alt)

        ws.freeze_panes = "A5"

    def _build_audit_sheet(self, wb, repo_name: str, now: datetime, config: dict):
        ws = wb.create_sheet("📝 Audit Trail")
        ws.sheet_view.showGridLines = False
        _add_title_block(ws, "Report Audit Trail", "Report generation metadata and configuration", now)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 60

        _header_row(ws, ["Attribute", "Value"], row=4)
        audit_data = [
            ("Report Generated At", now.strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Repository Name", repo_name),
            ("Report Format Version", "2.0"),
            ("Engine", "SENTINEL Reporting Engine v2.0"),
            ("Sheets Included", "9 (Summary, Health, Coverage, Delays, Drift, Violations, Trends, Commits, Audit)"),
            ("Configuration", str(config) if config else "Default"),
            ("Data Freshness", "Real-time at generation"),
        ]
        for idx, (attr, val) in enumerate(audit_data, start=5):
            _data_row(ws, [attr, val], row=idx, alt=idx % 2 == 0)

        ws.row_dimensions[4].height = 20
